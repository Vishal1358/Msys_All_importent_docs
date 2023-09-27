import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname" : "Rahul", "lastname" : "shetty", "gender":"Male"},{"firstname":"Anshika", "lastname" : "shetty", "gender":"Female"}]

    @staticmethod  # it is avoided the obj creations
    def getTestData(test_case_name):

        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\lenovo\\PycharmProjects1\\PythonSelFrame\\HomePageData\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]